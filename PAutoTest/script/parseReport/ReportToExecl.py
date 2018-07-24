#!/usr/bin/python3
# -*- coding: UTF-8 -*-

import sys
import re
import time
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
# import xlwt
# import xlrd 
# from xlutils.copy import copy 
from xml.dom.minidom import parse
import xml.dom.minidom


__CheckList_Sheet__= 'checkList'
__TEST_CASE_LINES__=4
__Result_Success__=r'success'
__Result_Failed__=r'failed'
__Case_result_Index__=1
__Case_name_Index__=0
caseTags=['case_name', 'case_result']
__Case_Elemate__=r'case'
__CheckList_First_Line__ = 24
__Check_List_Case_Name_Col__=3
import copy


def getNowTime():
    return time.strftime("%Y-%m-%d-%H-%M-%S",time.localtime(time.time()))


# class ExeclWriter:
# 	def __init__(self, filename):
# 		self.row = 0
# 		self.filename = filename
# 		self.book = xlwt.Workbook()
# 		self.sheet = self.book.add_sheet(getNowTime())
# 		col = 0
# 		for casetTag in caseTags :
# 			self.sheet.write(self.row, col, casetTag)
# 			col += 1
# 		self.row += 1	

# 	def write_end(self):
# 	    self.book.save(self.filename)

# 	def write_excel(self, reportElements):
# 		col = 0
# 		for reportElement in reportElements :
# 			self.sheet.write(self.row, col, reportElement)
# 			col += 1
# 		self.row += 1


def copyRange(srcSheet, srcRow, desSheet, desRow, maxCol) :
	for row in srcSheet.iter_rows(min_row=srcRow, max_row=srcRow + __TEST_CASE_LINES__ - 1, min_col=1, max_col=maxCol) :
		desCol = 2
		for cell in row :
			desCell = desSheet.cell(row = desRow, column = desCol)
			desCell.border = copy.copy(cell.border)
			desCell.value = cell.value
			# desCell.border = cell.border
			# print cell.value
			desCol += 1
		desRow += 1	


class ExeclChanger:
	def __init__(self, filename):
		self.row = 1
		self.filename = filename
		self.execl = openpyxl.load_workbook(self.filename)
		self.desSheet = self.execl.create_sheet(getNowTime())
		self.sourceSheet = self.execl[__CheckList_Sheet__]

	def getSrcRow(self, caseName) :
		maxRow = self.sourceSheet.max_row
		if __CheckList_First_Line__ > maxRow :
			return -1
		for row in range(__CheckList_First_Line__ + __TEST_CASE_LINES__ - 1, maxRow + 1, __TEST_CASE_LINES__) :
			curCaseName = self.sourceSheet.cell(row, __Check_List_Case_Name_Col__).value
			if curCaseName == caseName :
				return row + 1 - __TEST_CASE_LINES__
		return -1		

	def write_end(self):
	    self.execl.save(filename = self.filename)

	def write_excel(self, reportElements):
		if len(reportElements) != len(caseTags) :
			return

		caseName = reportElements[__Case_name_Index__]
		#self.desSheet.cell(row = self.row, column = col, value = caseName)
		caseRow = self.getSrcRow(caseName)
		if 	caseRow < 0 :
			print("caseRow error, ", caseName)
			return

		mergeCellStr='A' + str(self.row) + r':' + 'A' + str(self.row + __TEST_CASE_LINES__ - 1)
		self.desSheet.merge_cells(mergeCellStr)

		
		
		if reportElements[__Case_result_Index__] == __Result_Success__ :
			print "success"
			self.desSheet.cell(row = self.row, column = 1, value = '○')
		else:
			print "failed"
			self.desSheet.cell(row = self.row, column = 1, value = 'X')

		thin = Side(border_style="thin", color="FF000000")
		border = Border(top=thin, left=thin, right=thin, bottom=thin)
		self.desSheet['A' + str(self.row)].border = border
		copyRange(self.sourceSheet, caseRow, self.desSheet, self.row, self.sourceSheet.max_column)
		self.row += __TEST_CASE_LINES__



def trXMLToExecl(inputPath, outPutFile):
	DOMTree = xml.dom.minidom.parse(inputPath)
	rootElement = DOMTree.documentElement
	cases = rootElement.getElementsByTagName(__Case_Elemate__)
	execlWriter = ExeclChanger(outPutFile)
	for case in cases:
		reportElements = []
		for caseTag in caseTags:
			nodeList = case.getElementsByTagName(caseTag)
			if nodeList.length != 0 :
				node = nodeList[0]
				if node.firstChild != None :
					reportElements.append(node.firstChild.data)
				else :
					reportElements.append('-')
			else :
				reportElements.append('-')
		execlWriter.write_excel(reportElements)
	execlWriter.write_end()				

if __name__=='__main__':
	print(sys.argv)
	if len(sys.argv) >= 3 :
		file_path = sys.argv[1]
		outPut_file = sys.argv[2]
		print("out put result : ", outPut_file)
		trXMLToExecl(file_path, outPut_file)
	else :
		print("please input report and execl")
